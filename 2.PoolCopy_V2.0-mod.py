#encoding=utf-8
import openpyxl,os
from optparse import OptionParser
def DuanKouSaiXuan(_gp):
    if type(_gp) != type(1):
        Gp_list=_gp.strip(' ').split(' ')  #strip()移除字符串头尾指定的字符（默认为空格或换行符）或字符序列。 split()切片
    else:
        Gp_list=str(_gp)
    return Gp_list

def FW_snat(JB,JC,GIP,IIP):
    # fw_snat 传参GIP,SIP
    JB_SNAT_line=2
    JB['pool']['A2'].value=JC
    JB['pool']['B'+str(JB_SNAT_line)].value=GIP
    JB['pool']['D'+str(JB_SNAT_line)].value=IIP

def FW_SecurityPolicy(JB,IIP,SP,i):
    # Security-Policy 传参sip,sp
    JB_PSecurity_line=7+i
    JB['pool']['D'+str(JB_PSecurity_line)].value=IIP
    if '-' in str(SP):     #连续端口端判断
        JB['pool']['E'+str(JB_PSecurity_line)].value=SP.replace('-',' to ')
    else:
        JB['pool']['E'+str(JB_PSecurity_line)].value=SP
    JB_PSecurity_line+=1

def FW_dnat(JB_DNAT_line,JB,GIP,GP,IIP,SP):
    # NAT service 传参gip,gp,sip,sp
    if type(GP) != type(1): #端口数量判断，单端口筛选
        for i in range(GP.count(' ')+1):
            print (i,GIP,"|",DuanKouSaiXuan(GP)[i],"|",IIP,"|",DuanKouSaiXuan(SP)[i]) #映射端口检查
            print ('-----------------------------------------------------')
            if '-' in DuanKouSaiXuan(GP)[i]:#连续端口端判断
                JB['pool']['C'+str(JB_DNAT_line)].value=DuanKouSaiXuan(GP)[i].replace('-',' ')
                JB['pool']['E'+str(JB_DNAT_line)].value=DuanKouSaiXuan(SP)[i].replace('-',' ')
            else:
                JB['pool']['C'+str(JB_DNAT_line)].value=DuanKouSaiXuan(GP)[i]
                JB['pool']['E'+str(JB_DNAT_line)].value=DuanKouSaiXuan(SP)[i]
            JB['pool']['B'+str(JB_DNAT_line)].value=GIP
            JB['pool']['D'+str(JB_DNAT_line)].value=IIP
            JB_DNAT_line=JB_DNAT_line+1
    else:
         print (GIP,"|",DuanKouSaiXuan(GP),"|",IIP,"|",DuanKouSaiXuan(SP)) #映射端口检查
         print ('-----------------------------------------------------')
         JB['pool']['C'+str(JB_DNAT_line)].value=DuanKouSaiXuan(GP)
         JB['pool']['E'+str(JB_DNAT_line)].value=DuanKouSaiXuan(SP)
         JB['pool']['B'+str(JB_DNAT_line)].value=GIP
         JB['pool']['D'+str(JB_DNAT_line)].value=IIP
         JB_DNAT_line=JB_DNAT_line+1
    return JB_DNAT_line

def COPY_Run(JB,line,int_sheet,PJ_Number):
    JC=int_sheet['D'+str(line)].value
    print('NAME:',JC+'|','\n-------------------')
    JB_DNAT_line=17
    for i in range(PJ_Number):
        GIP=int_sheet['F'+str(line)].value
        GP=int_sheet['G'+str(line)].value
        IIP=int_sheet['H'+str(line)].value
        SP=int_sheet['I'+str(line)].value
        FIP=int_sheet['J'+str(line)].value
        FP=int_sheet['K'+str(line)].value
        line+=1
        if FIP != None:
            IIP=FIP
            SP=FP
        FW_snat(JB,JC,GIP,IIP)       #snat
        FW_SecurityPolicy(JB,IIP,SP,i) #Security-policy
        FW_dnat(JB_DNAT_line,JB,GIP,GP,IIP,SP)  #DNAT
        JB_DNAT_line=FW_dnat(JB_DNAT_line,JB,GIP,GP,IIP,SP)

def main():
    #--------------------------------------open_xlsx----------------------------
    fd_pool=openpyxl.load_workbook(r"service_IpPort_Open statistics.xlsx")
    JB=openpyxl.load_workbook(r"ConfigurationTemplate_V2.10.xlsx")

    #--------------------------------------定义option----------------------------
    optParser=OptionParser()
    # optParser.add_option("-l","--line",action="store",type="int",
        # dest="op_line",help='数据的起始行')
    # optParser.add_option("-n","--number",action="store",type="int",
        # dest="op_Number",help='操作多少条数据')
    optParser.add_option("-p","--pool",action="store",type="string",
        dest="op_pool",help='资源池选择，A网还是B网')
    options,args=optParser.parse_args()

    #-------------------------------------传参赋值-------------------------------------
    row=fd_pool[options.op_pool].max_row
    # column=fd_pool[options.op_pool].max_column
    for i in range(10):
        if fd_pool[options.op_pool]['B'+str(row-i)].value !=None:
            print('i的值为:'+str(i+1))
            line=row-i
            PJ_Number=i+1
            break
    # line=options.op_line
    # PJ_Number=options.op_Number
    int_sheet=fd_pool[options.op_pool]
    print('-----------------------------------------------------')
    print ('|第'+str(line)+'行|','|'+str(PJ_Number)+'行数据|','|资源池'+str(int_sheet)+'|')
    print('-----------------------------------------------------')

    #------------------------------------copy_Run-------------------------------------
    COPY_Run(JB,line,int_sheet,PJ_Number)
    JB.save(r'C:\Users\...\Desktop\cli_shell_Configuration.xlsx')
    os.system(r'C:\Users\...\Desktop\cli_shell_Configuration.xlsx')

if __name__ == "__main__":
    main()


